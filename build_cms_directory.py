"""
CMS Audit Provider Directory Consolidation
Extracts Year, CMS Category, Provider/Facility Name, NPI from:
  - AUDIT_CMS_PROVDIR_eVIP_20230105.xlsx  → 2023
  - AUDIT_CMS_PROVDIR_eVIP_20240104.xlsx  → 2024
  - AUDIT_CMS_PROVDIR_eVIP_20250102.xlsx  → 2025

CMS Category mapping per Johanna Tellechea / Dr. Alan Conrad / Cecy Villa:
  a. Primary Care – Adult
  b. Primary Care – Pediatric
  c. OB/GYN
  d. Mental Health – Adult
  e. Mental Health – Pediatric
  f. Specialists  (OB/GYN also appears here per confirmed guidance)
  g. Hospital
  h. Pharmacy
  i. Pediatric Dental  — OMITTED (no contracted providers per Johanna Tellechea)
  j. Long-Term Services and Supports  (SNF + RCF + CBAS, per Dr. Conrad & Cecy Villa)
"""

import pandas as pd
import os
from collections import Counter, defaultdict
from python_calamine import CalamineWorkbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILES = [
    (os.path.join(BASE_DIR, "AUDIT_CMS_PROVDIR_eVIP_20230105.xlsx"), 2023),
    (os.path.join(BASE_DIR, "AUDIT_CMS_PROVDIR_eVIP_20240104.xlsx"), 2024),
    (os.path.join(BASE_DIR, "AUDIT_CMS_PROVDIR_eVIP_20250102.xlsx"), 2025),
]
OUTPUT_FILE = os.path.join(BASE_DIR, "CMS_Audit_Provider_Directory_2023_2024_2025.xlsx")


# ── Category plan ─────────────────────────────────────────────────────────────
# (cms_label, [sheet_names], filter_fn | None)
# filter_fn receives a DataFrame with at least SPEC and AGE RESTRICTION columns
# and returns the filtered DataFrame.
CATEGORY_PLAN = [
    (
        "a. Primary Care – Adult",
        ["PROVDIR_PCP"],
        lambda df: df[
            ~df["SPEC"].str.upper().str.contains(r"\(PEDS\)", na=False, regex=True) &
            ~df["SPEC"].str.upper().str.strip().isin(["PEDIATRICS", "GENERAL PRACTICE (PEDS)"])
        ],
    ),
    (
        "b. Primary Care – Pediatric",
        ["PROVDIR_PCP"],
        lambda df: df[
            df["SPEC"].str.upper().str.contains(r"\(PEDS\)", na=False, regex=True) |
            df["SPEC"].str.upper().str.strip().isin(["PEDIATRICS", "GENERAL PRACTICE (PEDS)"])
        ],
    ),
    (
        "c. OB/GYN",
        ["PROVDIR_SPE"],
        lambda df: df[df["SPEC"].str.upper().str.strip() == "OBSTETRICS & GYNECOLOGY"],
    ),
    (
        "d. Mental Health – Adult",
        ["PROVDIR_MH"],
        lambda df: df[~(
            df["SPEC"].str.upper().str.contains("CHILD|ADOLESCENT", na=False, regex=True) |
            (
                df["SPEC"].str.upper().str.strip().eq("BEHAVIORAL HEALTH THERAPIES") &
                df["AGE RESTRICTION"].str.upper().str.contains(
                    r"UP TO|TO 18|TO 17|TO 6", na=False, regex=True
                )
            )
        )],
    ),
    (
        "e. Mental Health – Pediatric",
        ["PROVDIR_MH"],
        lambda df: df[
            df["SPEC"].str.upper().str.contains("CHILD|ADOLESCENT", na=False, regex=True) |
            (
                df["SPEC"].str.upper().str.strip().eq("BEHAVIORAL HEALTH THERAPIES") &
                df["AGE RESTRICTION"].str.upper().str.contains(
                    r"UP TO|TO 18|TO 17|TO 6", na=False, regex=True
                )
            )
        ],
    ),
    (
        "f. Specialists",
        ["PROVDIR_SPE"],
        None,  # all rows — OB/GYN included here too per confirmed guidance
    ),
    (
        "g. Hospital",
        ["PROVDIR_HOS", "PROVDIR_HOSGAH", "PROVDIR_HOSADD ", "PROVDIR_HOSMH"],
        None,
    ),
    (
        "h. Pharmacy",
        ["PROVDIR_PHARM"],
        None,
    ),
    # i. Pediatric Dental — omitted (no contracted providers)
    (
        "j. Long-Term Services and Supports",
        ["PROVDIR_SNF", "PROVDIR_RCF", "PROVDIR_CBAS"],
        None,
    ),
]

NEEDED_COLS = {"NPIis", "INDEXNM", "SPEC", "AGE RESTRICTION"}


def load_sheet(wb_path: str, sheet_name: str) -> pd.DataFrame | None:
    """Load only the needed columns from one sheet using the fast calamine engine."""
    try:
        df = pd.read_excel(
            wb_path,
            sheet_name=sheet_name,
            engine="calamine",
            dtype=str,
        )
        # Keep only the columns we need
        keep = [c for c in df.columns if c in NEEDED_COLS]
        df   = df[keep].copy()
        # Ensure all needed columns exist
        for col in NEEDED_COLS:
            if col not in df.columns:
                df[col] = ""
        df = df.fillna("")
        return df
    except Exception as e:
        return None


def main():
    all_records = []

    for wb_path, year in FILES:
        print(f"\n── {year}  ({os.path.basename(wb_path)}) ──")

        # Get actual sheet names (handles trailing spaces like 'PROVDIR_HOSADD ')
        cal_wb  = CalamineWorkbook.from_path(wb_path)
        avail   = {n.strip(): n for n in cal_wb.sheet_names}

        for cms_label, sheet_names, filter_fn in CATEGORY_PLAN:
            for raw_sheet in sheet_names:
                canonical = raw_sheet.strip()
                actual    = avail.get(canonical)
                if actual is None:
                    continue  # sheet not present in this year

                df = load_sheet(wb_path, actual)
                if df is None or df.empty:
                    continue

                # Apply category filter
                if filter_fn is not None:
                    try:
                        df = filter_fn(df)
                    except Exception as e:
                        print(f"  WARNING: filter error on {cms_label}/{actual}: {e}")
                        continue

                # Drop rows where NPI or Name is blank
                df = df[
                    (df["NPIis"].str.strip() != "") &
                    (df["INDEXNM"].str.strip() != "")
                ]

                n = len(df)
                print(f"  {cms_label:45s} | {canonical:20s} | {n:5,d} rows")

                for _, row in df.iterrows():
                    all_records.append((
                        year,
                        cms_label,
                        row["INDEXNM"].strip(),
                        row["NPIis"].strip(),
                    ))

    print(f"\nTotal rows before dedup : {len(all_records):,}")

    # De-duplicate by (year, CMS category, NPI)
    seen    = set()
    deduped = []
    for rec in all_records:
        key = (rec[0], rec[1], rec[3])   # year, category, npi
        if key not in seen:
            seen.add(key)
            deduped.append(rec)

    print(f"Total rows after  dedup : {len(deduped):,}")

    # Sort: Year → CMS Category → Provider Name
    deduped.sort(key=lambda r: (r[0], r[1], r[2]))

    # ── Write Excel ───────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    DARK_BLUE   = PatternFill("solid", fgColor="1F4E79")
    WHITE_BOLD  = Font(bold=True, color="FFFFFF", size=11)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    out_wb  = Workbook()
    ws_main = out_wb.active
    ws_main.title = "Provider Directory"

    ws_main.append(["Year", "CMS Category", "Provider or Facility Name", "NPI"])
    for cell in ws_main[1]:
        cell.font      = WHITE_BOLD
        cell.fill      = DARK_BLUE
        cell.alignment = CENTER

    for rec in deduped:
        ws_main.append(list(rec))

    for idx, width in enumerate([8, 42, 55, 14], 1):
        ws_main.column_dimensions[get_column_letter(idx)].width = width
    ws_main.freeze_panes = "A2"

    # Summary sheet
    ws_sum = out_wb.create_sheet("Summary")
    ws_sum.append(["Year", "CMS Category", "Provider Count (deduplicated)"])
    for cell in ws_sum[1]:
        cell.font      = WHITE_BOLD
        cell.fill      = DARK_BLUE
        cell.alignment = CENTER

    dedup_counts = Counter((r[0], r[1]) for r in deduped)
    for wb_path, year in FILES:
        for cms_label, *_ in CATEGORY_PLAN:
            ws_sum.append([year, cms_label, dedup_counts.get((year, cms_label), 0)])

    for idx, width in enumerate([8, 42, 28], 1):
        ws_sum.column_dimensions[get_column_letter(idx)].width = width
    ws_sum.freeze_panes = "A2"

    out_wb.save(OUTPUT_FILE)
    print(f"\n✓ Saved → {OUTPUT_FILE}")

    # Console summary
    print(f"\n{'Year':<6} {'CMS Category':<45} {'Count':>7}")
    print("-" * 62)
    for wb_path, year in FILES:
        for cms_label, *_ in CATEGORY_PLAN:
            print(f"{year:<6} {cms_label:<45} {dedup_counts.get((year, cms_label), 0):>7,}")
        print()


if __name__ == "__main__":
    main()
