# ==============================================================================
    # MODIFICATION(S) / CHANGE.LOG: 
# ==============================================================================

# Script to UPLOAD Excel file to SQL Server - optimized FOR "UPLOAD IPA Directory Data File_05.08.26.xlsx"
# DYNAMIC COLUMN READ ... 
# UPDATE (2026-01-29): Modified to upload IPA Provider Directory template to specific table
# UPDATE (2026-02-12): Modified file name, table name for February 2026 roster upload
# UPDATE (2026-03-12): Modified file name, table name for March 2026 roster upload
# UPDATE (2026-04-10): Modified file name, table name for April 2026 roster upload
# UPDATE (2026-05-11): Modified file name, table name for May 2026 roster upload
# CHARINDEX() FIND SEARCH() 'UPDATE HERE:' MARKERS TO MAKE NECESSARY CODE UPDATE(S)
# ALTER ADD [FIELD] REQUIRE(S) ... (#) UPDATE(S)

# ==============================================================================
# IMPORT LIBRARIES:
# ==============================================================================
import os
import glob
import pyodbc
from datetime import datetime
import openpyxl


def truncate_data(data, max_lengths):
    """Truncate data to fit column constraints"""
    return [str(item)[:length] if length and item else str(item) if item else None for item, length in zip(data, max_lengths)]


def get_column_max_lengths(cursor, table_name):
    """Get maximum character lengths for table columns"""
    cursor.execute(f"""
    SELECT COLUMN_NAME, CHARACTER_MAXIMUM_LENGTH
    FROM INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_NAME = '{table_name}'
    ORDER BY ORDINAL_POSITION
    """)
    return {row.COLUMN_NAME: row.CHARACTER_MAXIMUM_LENGTH for row in cursor.fetchall()}


def create_table(cursor, table_name, headers):
    """Create table structure DYNAMICally based on Excel headers"""
    column_count = len(headers)

    # Start building the CREATE TABLE statement
    sql_script = f"""
    USE INFORMATICS;

    IF OBJECT_ID('dbo.{table_name}', 'U') IS NOT NULL
        DROP TABLE dbo.{table_name};

    CREATE TABLE dbo.{table_name}
    (
        ID INT IDENTITY(1,1) PRIMARY KEY"""

    # Add columns based on actual Excel headers
    for i, header in enumerate(headers):
        # Clean header name for SQL column naming
        if header and str(header).strip():
            clean_header = str(header).strip()
            # Replace problematic characters for SQL column names
            clean_header = clean_header.replace(' ', '_').replace('-', '_').replace('/', '_')
            clean_header = clean_header.replace('(', '').replace(')', '').replace('.', '_')
            clean_header = clean_header.replace('#', 'NUM').replace('%', 'PCT').replace('&', 'AND')
            clean_header = clean_header.replace("'", '').replace('"', '').replace('`', '')
            # Remove any remaining special characters
            clean_header = ''.join(c for c in clean_header if c.isalnum() or c == '_')
            # Ensure it starts with a letter or underscore
            if clean_header and not (clean_header[0].isalpha() or clean_header[0] == '_'):
                clean_header = f'COL_{clean_header}'
            # Limit length to 128 characters (SQL Server limit)
            clean_header = clean_header[:128]
        else:
            clean_header = f'COL_{i+1}'

        # Determine appropriate data type based on header name - ENHANCED for IPA Provider Directory
        if any(keyword in clean_header.upper() for keyword in ['DATE', 'TIME', 'EFFECTIVE', 'TERM', 'EXPIR', 'RENEWAL']):
            data_type = 'DATETIME2'
        elif any(keyword in clean_header.upper() for keyword in ['NPI', 'ID', 'NUM', 'COUNT', 'AGE', 'LICENSE', 'TAX']):
            data_type = 'NVARCHAR(50)'  # Keep as string for flexibility
        elif any(keyword in clean_header.upper() for keyword in ['EMAIL', 'WEBSITE', 'URL', 'LINK', 'MAIL']):
            data_type = 'NVARCHAR(500)'
        elif any(keyword in clean_header.upper() for keyword in ['ADDRESS', 'DESC', 'DESCRIPTION', 'NOTES', 'SPECIALTY', 'CREDENTIAL', 'COMMENT']):
            data_type = 'NVARCHAR(1000)'
        elif any(keyword in clean_header.upper() for keyword in ['NAME', 'PROVIDER', 'ORGANIZATION', 'GROUP', 'PRACTICE', 'PPG', 'IPA', 'PHYSICIAN']):
            data_type = 'NVARCHAR(255)'
        elif any(keyword in clean_header.upper() for keyword in ['PHONE', 'ZIP', 'CODE', 'STATE', 'CITY', 'FAX', 'CONTACT']):
            data_type = 'NVARCHAR(50)'
        elif any(keyword in clean_header.upper() for keyword in ['STATUS', 'TYPE', 'CATEGORY', 'CLASS', 'IND', 'FLAG', 'TIER', 'NETWORK']):
            data_type = 'NVARCHAR(100)'
        else:
            data_type = 'NVARCHAR(255)'  # Default

        sql_script += f",\n        [{clean_header}] {data_type}"

    # Add audit columns
    sql_script += """,
        ImportedAt DATETIME2 DEFAULT GETDATE(),
        ImportedBy NVARCHAR(50) DEFAULT SYSTEM_USER
    );
    """

    # Execute the CREATE TABLE command
    commands = sql_script.split(';')

    for command in commands:
        if command.strip():
            try:
                cursor.execute(command)
            except pyodbc.Error as e:
                print(f"Error executing SQL command: {e}")
                print(f"Command: {command}")
                raise

    cursor.commit()
    print(f"Table {table_name} has been created with {column_count} data columns plus audit fields.")

    # Return the cleaned column names for later use
    cleaned_headers = []
    for i, header in enumerate(headers):
        if header and str(header).strip():
            clean_header = str(header).strip()
            clean_header = clean_header.replace(' ', '_').replace('-', '_').replace('/', '_')
            clean_header = clean_header.replace('(', '').replace(')', '').replace('.', '_')
            clean_header = clean_header.replace('#', 'NUM').replace('%', 'PCT').replace('&', 'AND')
            clean_header = clean_header.replace("'", '').replace('"', '').replace('`', '')
            clean_header = ''.join(c for c in clean_header if c.isalnum() or c == '_')
            if clean_header and not (clean_header[0].isalpha() or clean_header[0] == '_'):
                clean_header = f'COL_{clean_header}'
            clean_header = clean_header[:128]
        else:
            clean_header = f'COL_{i+1}'
        cleaned_headers.append(clean_header)

    return cleaned_headers


def stream_excel_to_sql(input_path, table_name, connection_string, batch_size=1000):
    """Main function to stream Excel data to SQL Server"""
    if not os.path.exists(input_path):
        raise ValueError(f"Input path does not exist: {input_path}")

    # UPDATE HERE: NEW FILE NAME for CHPIV ROSTER IPA Directory Data File - May 2026
    target_file = "UPLOAD IPA Directory Data File_05.08.26.xlsx"
    file_path = os.path.join(input_path, target_file)

    if not os.path.exists(file_path):
        # Also try looking with glob pattern
        all_files = glob.glob(os.path.join(input_path, target_file))
        if not all_files:
            # List available files for debugging
            available_files = [f for f in os.listdir(input_path) if f.endswith('.xlsx')]
            raise ValueError(f"Target file not found: {target_file}\nAvailable Excel files: {available_files}")
        file_path = all_files[0]

    try:
        conn = pyodbc.connect(connection_string, timeout=30)
        cursor = conn.cursor()

        start_time = datetime.now()
        processed_rows = 0

        print(f"Processing Excel file: {file_path}")

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active

        # Get total row count (excluding header)
        max_row = worksheet.max_row
        total_rows = max_row - 1  # Subtract header row
        print(f"Total data rows to process: {total_rows}")

        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value else "")

        print(f"Excel headers (first 10): {headers[:10]}")
        print(f"Total columns detected: {len(headers)}")

        # Create table with DYNAMIC structure based on headers
        cleaned_headers = create_table(cursor, table_name, headers)
        column_max_lengths = get_column_max_lengths(cursor, table_name)
        print("Table created and column max lengths retrieved successfully")

        # Create DYNAMIC INSERT statement
        columns_list = ', '.join([f'[{col}]' for col in cleaned_headers])
        placeholders = ','.join(['?'] * len(headers))

        insert_sql = f"""
        INSERT INTO {table_name} ({columns_list})
        VALUES ({placeholders})
        """

        print(f"Using {len(headers)} columns for INSERT statement")
        print("Processing Excel data rows...")

        # Get max lengths for truncation
        max_lengths = [column_max_lengths.get(col, None) for col in cleaned_headers]

        batch = []

        # Process data rows (skip header row)
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Convert Excel row to list and ensure we have exactly the right number of values
                row_data = list(row) if row else []

                # Adjust to match column count
                while len(row_data) < len(headers):
                    row_data.append(None)
                if len(row_data) > len(headers):
                    row_data = row_data[:len(headers)]

                # Clean the data - convert empty strings to None, strip whitespace
                processed_row = [
                    str(item).strip() if item is not None and str(item).strip() else None
                    for item in row_data
                ]

                # Apply max length truncation
                truncated_row = truncate_data(processed_row, max_lengths)
                batch.append(tuple(truncated_row))

                # Show progress for first few rows
                if row_num <= 5:
                    sample_data = [f"{cleaned_headers[i]}={truncated_row[i]}" for i in range(min(3, len(truncated_row)))]
                    print(f"Sample row {row_num}: {', '.join(sample_data)}")

                # Insert batch when it reaches batch_size
                if len(batch) >= batch_size:
                    try:
                        cursor.executemany(insert_sql, batch)
                        conn.commit()
                        processed_rows += len(batch)
                        print(f"Progress: {processed_rows}/{total_rows} rows ({processed_rows/total_rows*100:.1f}%)")
                        batch = []
                    except pyodbc.Error as e:
                        print(f"Error inserting batch: {e}")
                        # Try inserting rows one by one to identify problematic rows
                        for i, single_row in enumerate(batch):
                            try:
                                cursor.execute(insert_sql, single_row)
                                conn.commit()
                                processed_rows += 1
                            except pyodbc.Error as row_error:
                                print(f"Error inserting row {row_num - len(batch) + i}: {row_error}")
                        batch = []

            except Exception as e:
                print(f"Error processing row {row_num}: {e}")
                continue

        # Insert remaining batch
        if batch:
            try:
                cursor.executemany(insert_sql, batch)
                conn.commit()
                processed_rows += len(batch)
            except pyodbc.Error as e:
                print(f"Error inserting final batch: {e}")
                for i, single_row in enumerate(batch):
                    try:
                        cursor.execute(insert_sql, single_row)
                        conn.commit()
                        processed_rows += 1
                    except pyodbc.Error as row_error:
                        print(f"Error inserting final row: {row_error}")

        workbook.close()
        print(f"Finished processing: {file_path}")

        end_time = datetime.now()
        duration = end_time - start_time
        print(f"All data has been streamed to SQL database.")
        print(f"Total rows processed: {processed_rows}")
        print(f"Time taken: {duration}")

        # Validation query to verify data upload
        cursor.execute(f"SELECT COUNT(*) as RecordCount FROM {table_name}")
        final_count = cursor.fetchone().RecordCount
        print(f"Final record count in database: {final_count}")

        # Display sample of uploaded data
        cursor.execute(f"SELECT TOP 5 * FROM {table_name} ORDER BY ImportedAt DESC")
        sample_rows = cursor.fetchall()
        print("\nSample of uploaded data:")
        for row in sample_rows:
            print(f"ID: {row[0]}, First few columns: {list(row[1:4])}")

    except Exception as e:
        print(f"An error occurred: {e}")
        raise
    finally:
        if 'conn' in locals():
            conn.close()

# ==============================================================================
    # Configuration - UPDATE HERE: CHPIV ROSTER IPA Directory Data File specifications - May 2026
# ==============================================================================
input_path = r"C:\Users\wcarr\Desktop"

# UPDATE HERE: Fixed table name as specified - May 2026 update
table_name = "CHPIV_PROVIDER_DIRECTORY_DATA_UPDATE_20260511"

# UPDATE HERE: Connection string pointing to SQLPRODAPP01
connection_string = "DRIVER={SQL Server};SERVER=SQLPRODAPP01;DATABASE=INFORMATICS;Trusted_Connection=yes;"


# Execute the function
if __name__ == "__main__":
    print("=" * 100)
    print("CHPIV ROSTER IPA PROVIDER DIRECTORY UPLOAD SCRIPT - MAY 2026 UPDATE")
    print("UPLOAD IPA Directory Data File_05.08.26.xlsx to SQL Server")
    print("=" * 100)
    print(f"Source File: {os.path.join(input_path, 'UPLOAD IPA Directory Data File_05.08.26.xlsx')}")
    print(f"Target Table: SQLPRODAPP01.INFORMATICS.dbo.{table_name}")
    print("=" * 100)

    try:
        stream_excel_to_sql(input_path, table_name, connection_string)
        print("\n" + "=" * 100)
        print("✅ UPLOAD COMPLETED SUCCESSFULLY!")
        print("=" * 100)
    except Exception as e:
        print("\n" + "=" * 100)
        print("❌ UPLOAD FAILED!")
        print(f"Error: {e}")
        print("=" * 100)
